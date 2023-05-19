from openpyxl import load_workbook
from django.shortcuts import render, redirect

from app01 import models
from app01.utils.pagination import Pagination


# Create your views here.
def depart_list(request):
    """ 部门列表 """

    # 去数据库中获取所有的部门列表
    # [对象,对象,对象]
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset)
    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.html
    }
    return render(request, "depart_list.html", context)


def depart_add(request):
    """ 添加部门 """
    if request.method == "GET":
        return render(request, "depart_add.html")

    # 获取用户POST提交过来的数据
    title = request.POST.get("title")

    # 保存到数据库
    models.Department.objects.create(title=title)

    # 重定向回部门列表
    return redirect("/depart/list/")


def depart_delete(request):
    """ 删除部门 """
    # 获取ID http://127.0.0.1:8000/depart/delete/?nid=1
    nid = request.GET.get('nid')

    # 删除
    models.Department.objects.filter(id=nid).delete()

    # 重定向回部门列表
    return redirect("/depart/list/")


def depart_edit(request, nid):
    """ 修改部门 """
    if request.method == "GET":
        # 根据nid，获取他的数据
        row_obj = models.Department.objects.filter(id=nid).first()

        return render(request, "depart_edit.html", {"row_obj": row_obj})

    # 获取用户提交的标题
    title = request.POST.get("title")

    # 根据ID找到数据库中的数据并进行更新
    models.Department.objects.filter(id=nid).update(title=title)

    # 重定向回部门列表
    return redirect("/depart/list/")


def depart_multi(request):
    """ 批量上传（Excel文件） """

    # 1. 获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2. 对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3. 循环换取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')